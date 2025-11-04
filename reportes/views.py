from django.shortcuts import render
from django.db.models import Sum, F, Count
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import datetime

# --- Importaciones de Excel ---
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# --- Importaciones de PDF ---
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# --- Importar modelo ---
from dashboard.models import Product


# === VISTA PRINCIPAL DE REPORTES ===
@login_required
def reportes(request):
    NIVEL_CRITICO = 5
    NIVEL_BAJO = 10

    # 🔒 Filtrar solo productos del usuario
    productos_criticos = Product.objects.filter(created_by=request.user, stock__lte=NIVEL_CRITICO)
    productos_bajo_stock = Product.objects.filter(created_by=request.user, stock__gt=NIVEL_CRITICO, stock__lte=NIVEL_BAJO)
    lista_reporte_stock = Product.objects.filter(created_by=request.user, stock__lte=NIVEL_BAJO).order_by('stock')

    count_criticos = productos_criticos.count()
    count_bajo_stock = productos_bajo_stock.count()
    valor_en_riesgo = productos_criticos.aggregate(total_valor=Sum(F('stock') * F('price')))['total_valor'] or 0
    total_categorias = Product.objects.filter(created_by=request.user).values('category').distinct().count()

    TARGET_STOCK = 15
    reorden_sug_val = 0
    for p in lista_reporte_stock:
        if p.stock < TARGET_STOCK:
            unidades_a_pedir = TARGET_STOCK - p.stock
            reorden_sug_val += unidades_a_pedir * p.price

    stock_total_inventario_unidades = Product.objects.filter(created_by=request.user).aggregate(
        total_stock=Sum('stock')
    )['total_stock'] or 1

    analisis_categorias = Product.objects.filter(created_by=request.user).values('category').annotate(
        num_productos=Count('id'),
        stock_total_categoria=Sum('stock'),
        valor_total_categoria=Sum(F('stock') * F('price'))
    ).order_by('-stock_total_categoria')

    for cat in analisis_categorias:
        stock_cat = cat['stock_total_categoria'] or 0
        cat['porcentaje'] = (stock_cat / stock_total_inventario_unidades) * 100
        if cat['valor_total_categoria'] is None:
            cat['valor_total_categoria'] = 0

    context = {
        'count_criticos': count_criticos,
        'count_bajo_stock': count_bajo_stock,
        'valor_en_riesgo': valor_en_riesgo,
        'total_categorias': total_categorias,
        'lista_reporte_stock': lista_reporte_stock,
        'impacto_reorden_sugerido': reorden_sug_val,
        'NIVEL_CRITICO': NIVEL_CRITICO,
        'NIVEL_BAJO': NIVEL_BAJO,
        'analisis_categorias': analisis_categorias,
    }

    return render(request, "reportes/reporte.html", context)


# === EXPORTAR EXCEL COMPLETO ===
@login_required
def exportar_excel_completo(request):
    NIVEL_CRITICO = 5
    NIVEL_BAJO = 10

    productos_criticos = Product.objects.filter(created_by=request.user, stock__lte=NIVEL_CRITICO)
    productos_bajo_stock = Product.objects.filter(created_by=request.user, stock__gt=NIVEL_CRITICO, stock__lte=NIVEL_BAJO)

    count_criticos = productos_criticos.count()
    count_bajo_stock = productos_bajo_stock.count()
    valor_en_riesgo = productos_criticos.aggregate(total_valor=Sum(F('stock') * F('price')))['total_valor'] or 0
    total_categorias = Product.objects.filter(created_by=request.user).values('category').distinct().count()
    stock_total_inventario_unidades = Product.objects.filter(created_by=request.user).aggregate(total_stock=Sum('stock'))['total_stock'] or 1

    analisis_categorias = Product.objects.filter(created_by=request.user).values('category').annotate(
        num_productos=Count('id'),
        stock_total_categoria=Sum('stock'),
        valor_total_categoria=Sum(F('stock') * F('price'))
    ).order_by('-valor_total_categoria')

    fecha_actual = datetime.datetime.now().strftime('%Y-%m-%d')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = f'attachment; filename="Reporte_Inventario_{fecha_actual}.xlsx"'

    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen de Inventario"

    font_titulo = Font(bold=True, size=16)
    font_subtitulo = Font(bold=True, size=12)
    font_bold = Font(bold=True)
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    align_center = Alignment(horizontal='center')
    align_right = Alignment(horizontal='right')
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws_resumen['B2'] = "Reporte de Inventario"; ws_resumen['B2'].font = font_titulo
    ws_resumen['B3'] = f"Generado el: {fecha_actual}"
    ws_resumen['B5'] = "Indicadores Clave"; ws_resumen['B5'].font = font_subtitulo
    ws_resumen['B6'] = "Productos Críticos (<= 5 un.)"
    ws_resumen['B7'] = "Bajo Stock (<= 10 un.)"
    ws_resumen['B8'] = "Valor en Riesgo (Críticos)"
    ws_resumen['B9'] = "Total de Categorías"

    ws_resumen['C6'] = count_criticos
    ws_resumen['C7'] = count_bajo_stock
    ws_resumen['C8'] = valor_en_riesgo
    ws_resumen['C9'] = total_categorias

    for row in ws_resumen['B6':'C9']:
        for cell in row:
            if cell.column_letter == 'B':
                cell.font = font_bold
            if cell.column_letter == 'C':
                cell.alignment = align_right
            cell.border = border_thin

    ws_resumen['C8'].number_format = '$ #,##0.00'

    ws_resumen['E5'] = "Análisis por Categoría"; ws_resumen['E5'].font = font_subtitulo
    headers_cat = ['Categoría', 'N° Productos', 'Total Unidades', '% Stock Total', 'Valor Total']
    for col_num, header in enumerate(headers_cat, 5):
        cell = ws_resumen.cell(row=6, column=col_num, value=header)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    row_num_cat = 7
    for cat in analisis_categorias:
        ws_resumen.cell(row=row_num_cat, column=5, value=cat['category'])
        ws_resumen.cell(row=row_num_cat, column=6, value=cat['num_productos'])
        ws_resumen.cell(row=row_num_cat, column=7, value=cat['stock_total_categoria'])
        ws_resumen.cell(row=row_num_cat, column=8, value=(cat['stock_total_categoria'] or 0) / stock_total_inventario_unidades * 100)
        ws_resumen.cell(row=row_num_cat, column=9, value=cat['valor_total_categoria'])
        row_num_cat += 1

    # === Hoja de Datos ===
    ws_datos = wb.create_sheet(title="Datos Completos")
    columnas_datos = ['ID', 'Nombre', 'Categoría', 'Stock', 'Precio Unitario', 'Valor Total Inventario', 'Estado']
    for col_num, column_title in enumerate(columnas_datos, 1):
        cell = ws_datos.cell(row=1, column=col_num, value=column_title)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center

    queryset = Product.objects.filter(created_by=request.user)
    row_num = 2
    for producto in queryset:
        valor_total = producto.stock * producto.price
        datos_fila = [producto.id, producto.name, producto.category,
                      producto.stock, producto.price, valor_total, producto.get_status_display()]
        for col_num, value in enumerate(datos_fila, 1):
            cell = ws_datos.cell(row=row_num, column=col_num, value=value)
        row_num += 1

    wb.save(response)
    return response


# === EXPORTAR EXCEL STOCK BAJO ===
@login_required
def exportar_excel(request):
    NIVEL_BAJO = 10
    lista_reporte_stock = Product.objects.filter(created_by=request.user, stock__lte=NIVEL_BAJO).order_by('stock')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    fecha_actual = datetime.datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Stock_Bajo_{fecha_actual}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Bajo"

    headers = ['ID', 'Nombre', 'Categoría', 'Stock Actual', 'Precio Unitario', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    row_num = 2
    for p in lista_reporte_stock:
        datos_fila = [p.id, p.name, p.category, p.stock, p.price, p.get_status_display()]
        for col_num, val in enumerate(datos_fila, 1):
            ws.cell(row=row_num, column=col_num, value=val)
        row_num += 1

    wb.save(response)
    return response


# === FUNCIONES PDF ===
def _get_pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TitlePDF', fontSize=18, alignment=1, spaceAfter=20))
    styles.add(ParagraphStyle(name='Heading2PDF', fontSize=14, fontName='Helvetica-Bold', spaceAfter=12, spaceBefore=12))
    styles.add(ParagraphStyle(name='BodyPDF', fontSize=10, spaceAfter=6))
    return styles


def _get_pdf_table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])


# === EXPORTAR PDF COMPLETO ===
@login_required
def exportar_pdf_completo(request):
    response = HttpResponse(content_type='application/pdf')
    fecha_actual = datetime.datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Inventario_Completo_{fecha_actual}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    story = []
    styles = _get_pdf_styles()

    story.append(Paragraph("Reporte de Inventario Completo", styles['TitlePDF']))
    story.append(Paragraph(f"Generado el: {fecha_actual}", styles['BodyPDF']))
    story.append(Spacer(1, 0.25 * inch))

    story.append(Paragraph("Análisis por Categoría", styles['Heading2PDF']))
    analisis_categorias = Product.objects.filter(created_by=request.user).values('category').annotate(
        num_productos=Count('id'),
        stock_total_categoria=Sum('stock'),
        valor_total_categoria=Sum(F('stock') * F('price'))
    ).order_by('-valor_total_categoria')

    data_cat = [['Categoría', 'N° Productos', 'Total Unidades', 'Valor Total']]
    for cat in analisis_categorias:
        data_cat.append([
            cat['category'], cat['num_productos'], cat['stock_total_categoria'],
            f"$ {cat['valor_total_categoria'] or 0:,.2f}"
        ])

    table_cat = Table(data_cat, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 2*inch])
    table_cat.setStyle(_get_pdf_table_style())
    story.append(table_cat)
    story.append(PageBreak())

    story.append(Paragraph("Inventario Completo (Todos los Productos)", styles['Heading2PDF']))
    productos = Product.objects.filter(created_by=request.user)
    data_prod = [['ID', 'Nombre', 'Categoría', 'Stock', 'Precio', 'Valor Total', 'Estado']]
    for p in productos:
        data_prod.append([
            p.id, Paragraph(p.name, styles['BodyPDF']), p.category, p.stock,
            f"$ {p.price:,.2f}", f"$ {p.stock * p.price:,.2f}", p.get_status_display()
        ])

    table_prod = Table(data_prod, colWidths=[0.5*inch, 3*inch, 1.5*inch, 0.75*inch, 1.25*inch, 1.5*inch, 1.25*inch])
    table_prod.setStyle(_get_pdf_table_style())
    story.append(table_prod)

    doc.build(story)
    return response


# === EXPORTAR PDF STOCK BAJO ===
@login_required
def exportar_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    fecha_actual = datetime.datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Stock_Bajo_{fecha_actual}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    story = []
    styles = _get_pdf_styles()

    story.append(Paragraph("Reporte de Stock Bajo", styles['TitlePDF']))
    story.append(Paragraph(f"Generado el: {fecha_actual}", styles['BodyPDF']))
    story.append(Spacer(1, 0.25 * inch))

    NIVEL_BAJO = 10
    lista_reporte_stock = Product.objects.filter(created_by=request.user, stock__lte=NIVEL_BAJO).order_by('stock')
    story.append(Paragraph(f"Productos con {NIVEL_BAJO} unidades o menos", styles['Heading2PDF']))

    data_stock = [['ID', 'Nombre', 'Categoría', 'Stock Actual', 'Estado']]
    for p in lista_reporte_stock:
        data_stock.append([
            p.id, Paragraph(p.name, styles['BodyPDF']), p.category, p.stock, p.get_status_display()
        ])

    table_stock = Table(data_stock, colWidths=[0.5*inch, 3*inch, 1.5*inch, 1*inch, 1*inch])
    table_stock.setStyle(_get_pdf_table_style())
    story.append(table_stock)

    productos_criticos = Product.objects.filter(created_by=request.user, stock__lte=5)
    count_criticos = productos_criticos.count()
    valor_en_riesgo = productos_criticos.aggregate(total_valor=Sum(F('stock') * F('price')))['total_valor'] or 0

    story.append(Spacer(1, 0.25 * inch))
    story.append(Paragraph("Impacto Financiero (Productos Críticos)", styles['Heading2PDF']))

    data_impacto = [
        ['Productos en riesgo (Críticos)', count_criticos],
        ['Valor total en riesgo', f"$ {valor_en_riesgo:,.2f}"]
    ]
    table_impacto = Table(data_impacto, colWidths=[3.5*inch, 3.5*inch])
    table_impacto.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    story.append(table_impacto)

    doc.build(story)
    return response
